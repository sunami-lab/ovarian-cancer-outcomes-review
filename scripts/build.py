#!/usr/bin/env python3
"""Build the evidence-extraction workbook and per-category Markdown from data/*.yml.

Usage: python3 scripts/build.py
Outputs: output/search_strategies.xlsx  (one sheet per category)
         output/search_strategies.md    (browsable mirror of the workbook, for GitHub)
         output/csv/<id>_<name>.csv     (per sheet; GitHub renders CSV as a live table)
         docs/tables/<id>_<name>.md     (per category, one section per review)

Everything is generated from data/NN_*.yml, so the four views cannot drift apart.
"""
import os
import re
import sys

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data")
OUT_XLSX = os.path.join(ROOT, "output", "search_strategies.xlsx")
OUT_MD = os.path.join(ROOT, "docs", "tables")

# (spreadsheet header, key in the YAML record, column width)
ANNEX_COLUMNS = [
    ("Title", "title", 60),
    ("1st author / year", "first_author_year", 18),
    ("Journal", "journal", 30),
    ("Impact factor", "impact_factor", 16),
    ("Why it is relevant to this review", "why_relevant", 62),
    ("Why it was excluded", "why_excluded", 34),
    ("URL", "url", 42),
]

COLUMNS = [
    ("Title", "title", 46),
    ("1st author / year", "first_author_year", 18),
    ("Journal", "journal", 26),
    ("Impact factor", "impact_factor", 14),
    ("Purpose of the study", "purpose", 52),
    ("P - Population", "pico_p", 32),
    ("I - Intervention / exposure", "pico_i", 32),
    ("C - Comparison", "pico_c", 32),
    ("O - Outcome", "pico_o", 32),
    ("Search terms / strings - PubMed / MEDLINE", "search_pubmed", 62),
    ("Search terms / strings - Embase", "search_embase", 62),
    ("Search terms / strings - Web of Science", "search_wos", 52),
    ("URL", "url", 40),
]

# One typeface throughout the workbook: Arial 11.
FONT_NAME = "Arial"
FONT_SIZE = 11

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=FONT_SIZE, name=FONT_NAME)
BODY_FONT = Font(size=FONT_SIZE, name=FONT_NAME)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BAND = PatternFill("solid", fgColor="F2F5FA")


def load():
    """Category files are named NN_<slug>.yml; anything else in data/ is ignored."""
    cats = []
    for fn in sorted(os.listdir(DATA)):
        if not re.match(r"\d{2}_.*\.ya?ml$", fn):
            continue
        with open(os.path.join(DATA, fn)) as fh:
            cats.append(yaml.safe_load(fh))
    return cats


def clean(v):
    """YAML folded scalars end with a newline; strip trailing whitespace only."""
    return (v or "").rstrip() if isinstance(v, str) else ("" if v is None else str(v))


# Excel's hard maximum row height is 409.5 points. Asking for more is silently clamped,
# which is why very long strategy cells cannot be fully displayed in the grid however the
# row is sized. The cell CONTENT is complete either way - it is visible in the formula bar,
# and rendered in full in docs/tables/*.md and strategies/*.txt.
EXCEL_MAX_ROW_HEIGHT = 409.0


def est_height(rec, cols=None):
    """Row height that shows as much of the tallest cell as Excel permits."""
    lines = 1
    for header, key, width in (cols or COLUMNS):
        txt = clean(rec.get(key))
        n = sum(max(1, -(-len(ln) // max(width - 2, 10))) for ln in txt.split("\n"))
        lines = max(lines, n)
    return min(max(lines, 4) * (FONT_SIZE * 1.32), EXCEL_MAX_ROW_HEIGHT)


def build_xlsx(cats):
    wb = Workbook()
    wb.remove(wb.active)
    for cat in cats:
        meta = cat["category"]
        cols = ANNEX_COLUMNS if meta.get("column_set") == "annex" else COLUMNS
        ws = wb.create_sheet(meta["sheet_name"][:31])
        ws.freeze_panes = "A3"

        ws.cell(row=1, column=1, value=f"{meta['id']}. {meta['title']}").font = Font(
            bold=True, size=FONT_SIZE, name=FONT_NAME)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.row_dimensions[1].height = 20

        for i, (header, key, width) in enumerate(cols, start=1):
            c = ws.cell(row=2, column=i, value=header)
            c.fill, c.font, c.border = HEAD_FILL, HEAD_FONT, BORDER
            c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[2].height = 30

        overflow = []
        for r, rec in enumerate(cat["records"], start=3):
            for i, (header, key, width) in enumerate(cols, start=1):
                c = ws.cell(row=r, column=i, value=clean(rec.get(key)))
                c.font = BODY_FONT
                c.alignment = Alignment(vertical="top", wrap_text=True)
                c.border = BORDER
                if r % 2:
                    c.fill = BAND
            h = est_height(rec, cols)
            ws.row_dimensions[r].height = h
            if h >= EXCEL_MAX_ROW_HEIGHT:
                overflow.append(r)
        ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{2 + len(cat['records'])}"
        if overflow:
            n = 3 + len(cat["records"])
            c = ws.cell(row=n, column=1, value=(
                "Rows " + ", ".join(map(str, overflow)) + ": some cells are longer than "
                "Excel's maximum row height (409.5 pt) allows it to display. The cell "
                "content is complete - click the cell and read the formula bar, or use "
                "docs/tables/*.md and strategies/*.txt for the full text."))
            c.font = Font(size=FONT_SIZE, italic=True, name=FONT_NAME, color="7F7F7F")
            c.alignment = Alignment(vertical="top", wrap_text=True)
            ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=len(cols))
            ws.row_dimensions[n].height = 28

    os.makedirs(os.path.dirname(OUT_XLSX), exist_ok=True)
    wb.save(OUT_XLSX)
    return OUT_XLSX


def md_cell(v):
    return clean(v).replace("|", "\\|").replace("\n", "<br>")


def build_companion(cats):
    """One browsable Markdown mirror of the whole workbook, for reading on GitHub.

    Generated from the same YAML as the .xlsx, so the two cannot drift. The three
    search-string columns hold up to 3 KB each; putting them inline would make the table
    unreadable, so each sheet gets a table of the short columns followed by the full
    strategies in collapsible <details> blocks, which GitHub renders natively.
    """
    path = os.path.join(ROOT, "output", "search_strategies.md")
    n_rec = sum(len(c["records"]) for c in cats if c["category"].get("column_set") != "annex")
    L = ["# Search strategies — browsable companion to `search_strategies.xlsx`", "",
         "Same content as the workbook, generated from the same source (`data/NN_*.yml`) by",
         "`scripts/build.py`, so the two cannot drift apart. Here so you can read the tables on",
         "GitHub without downloading the spreadsheet.", "",
         f"**{n_rec} systematic reviews across 8 categories, plus an annex of 10 reviews "
         "excluded by the impact-factor rule.**", "",
         "Each sheet below gives a table of the descriptive columns, then the full verbatim",
         "search strategies in collapsible blocks — click to expand. Where a review did not",
         "search a database, or searched it without publishing the string, the cell says so.", "",
         "## Contents", ""]
    for cat in cats:
        m = cat["category"]
        L.append(f"- [{m['id']}. {m['title']}](#{anchor(m)})")
    L.append("")

    for cat in cats:
        meta = cat["category"]
        L += ["---", "", f"## {meta['id']}. {meta['title']}", ""]
        L += [f"**Focus.** {clean(meta['review_question_focus'])}", ""]
        L += [f"**Population scope applied.** {clean(meta['population_scope'])}", ""]
        if meta.get("cross_cutting_note"):
            L += [f"**Note.** {clean(meta['cross_cutting_note'])}", ""]

        if meta.get("column_set") == "annex":
            L += ["| Title | 1st author / year | Journal | Impact factor | Why relevant | Why excluded | URL |",
                  "|---|---|---|---|---|---|---|"]
            for rec in cat["records"]:
                L.append("| " + " | ".join(md_cell(rec.get(k)) for k in
                         ("title", "first_author_year", "journal", "impact_factor",
                          "why_relevant", "why_excluded", "url")) + " |")
            L.append("")
            continue

        L += ["| Title | 1st author / year | Journal | Impact factor | Purpose of the study "
              "| P — Population | I — Intervention / exposure | C — Comparison | O — Outcome | URL |",
              "|---|---|---|---|---|---|---|---|---|---|"]
        for rec in cat["records"]:
            L.append("| " + " | ".join(md_cell(rec.get(k)) for k in
                     ("title", "first_author_year", "journal", "impact_factor", "purpose",
                      "pico_p", "pico_i", "pico_c", "pico_o", "url")) + " |")
        L += ["", "### Search strategies", ""]
        for rec in cat["records"]:
            L += [f"<details>", f"<summary><b>{clean(rec['first_author_year'])}</b> — "
                  f"{clean(rec['journal'])}</summary>", ""]
            if rec.get("strategy_source"):
                L += [f"Taken from: {clean(rec['strategy_source'])}", ""]
            for label, key in (("PubMed / MEDLINE", "search_pubmed"),
                               ("Embase", "search_embase"),
                               ("Web of Science", "search_wos")):
                L += [f"**{label}**", "", "```", clean(rec[key]), "```", ""]
            L += ["</details>", ""]

    open(path, "w").write("\n".join(L) + "\n")
    return path


def anchor(meta):
    """GitHub heading anchor for '## <id>. <title>'."""
    t = f"{meta['id']}-{meta['title']}".lower()
    return re.sub(r"[^a-z0-9\s-]", "", t).replace(" ", "-")


def build_csv(cats):
    """Per-sheet CSV. GitHub renders .csv with an interactive, searchable table viewer,
    which is the closest thing to opening the workbook in the browser."""
    import csv
    outdir = os.path.join(ROOT, "output", "csv")
    os.makedirs(outdir, exist_ok=True)
    written = []
    for cat in cats:
        meta = cat["category"]
        cols = ANNEX_COLUMNS if meta.get("column_set") == "annex" else COLUMNS
        slug = re.sub(r"[^a-z0-9]+", "_", meta["title"].lower()).strip("_")[:60]
        p = os.path.join(outdir, f"{meta['id']}_{slug}.csv")
        with open(p, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow([h for h, _, _ in cols])
            for rec in cat["records"]:
                w.writerow([clean(rec.get(k)) for _, k, _ in cols])
        written.append(p)
    return written


def build_md(cats):
    os.makedirs(OUT_MD, exist_ok=True)
    written = []
    for cat in cats:
        meta = cat["category"]
        slug = re.sub(r"[^a-z0-9]+", "_", meta["title"].lower()).strip("_")
        path = os.path.join(OUT_MD, f"{meta['id']}_{slug}.md")
        L = [f"# {meta['id']}. {meta['title']}", ""]
        L += [f"**Focus of this category.** {clean(meta['review_question_focus'])}", ""]
        L += [f"**Population scope applied.** {clean(meta['population_scope'])}", ""]
        if meta.get("column_set") == "annex":
            L += ["| Title | 1st author / year | Journal | Impact factor | Why relevant | Why excluded | URL |",
                  "|---|---|---|---|---|---|---|"]
            for rec in cat["records"]:
                L.append("| " + " | ".join(md_cell(rec.get(k)) for k in
                         ("title", "first_author_year", "journal", "impact_factor",
                          "why_relevant", "why_excluded", "url")) + " |")
            open(path, "w").write("\n".join(L) + "\n")
            written.append(path)
            continue
        for rec in cat["records"]:
            L += [f"## {clean(rec['first_author_year'])} - {clean(rec['journal'])}", ""]
            L += [f"**{clean(rec['title'])}**", ""]
            L += [f"- Impact factor: {clean(rec['impact_factor'])}"]
            L += [f"- URL: <{clean(rec['url'])}>"]
            if rec.get("pmid"):
                L += [f"- PMID: {clean(rec['pmid'])}" + (
                    f" | PMCID: {clean(rec['pmcid'])}" if rec.get("pmcid") else "")]
            if rec.get("strategy_source"):
                L += [f"- Search strategy taken from: {clean(rec['strategy_source'])}"]
            L += ["", "**Purpose.** " + clean(rec["purpose"]), ""]
            L += ["**PICO**", "",
                  "| | |", "|---|---|",
                  f"| **P** — Population | {md_cell(rec['pico_p'])} |",
                  f"| **I** — Intervention / exposure | {md_cell(rec['pico_i'])} |",
                  f"| **C** — Comparison | {md_cell(rec['pico_c'])} |",
                  f"| **O** — Outcome | {md_cell(rec['pico_o'])} |", ""]
            for label, key in (("PubMed / MEDLINE", "search_pubmed"),
                               ("Embase", "search_embase"),
                               ("Web of Science", "search_wos")):
                L += [f"**Search strategy - {label}**", "", "```",
                      clean(rec[key]), "```", ""]
        open(path, "w").write("\n".join(L) + "\n")
        written.append(path)
    return written


def main():
    cats = load()
    if not cats:
        sys.exit("No YAML files found in data/")
    x = build_xlsx(cats)
    m = build_md(cats)
    comp = build_companion(cats)
    csvs = build_csv(cats)
    n = sum(len(c["records"]) for c in cats)
    print(f"{len(cats)} categories, {n} reviews")
    print("wrote", os.path.relpath(x, ROOT))
    print("wrote", os.path.relpath(comp, ROOT))
    print(f"wrote {len(csvs)} files in output/csv/")
    for p in m:
        print("wrote", os.path.relpath(p, ROOT))


if __name__ == "__main__":
    main()
